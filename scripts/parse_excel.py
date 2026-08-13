import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def norm(value):
    if value is None:
        return ""
    if hasattr(value, "date"):
        try:
            return value.isoformat()
        except Exception:
            return str(value)
    return str(value).strip()


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Usage: parse_excel.py <workbook.xlsx>"}))
        sys.exit(1)

    path = Path(sys.argv[1])
    if not path.exists():
        print(json.dumps({"error": f"Workbook not found: {path}"}))
        sys.exit(1)

    wb = load_workbook(path, data_only=True)
    entries = []
    codes = {}

    if "Codes" in wb.sheetnames:
        ws = wb["Codes"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            headers = [norm(h).lower() for h in rows[0]]
            for row in rows[1:]:
                item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
                name = norm(item.get("project_name") or item.get("name"))
                code = norm(item.get("code") or item.get("internal_order_code"))
                if name and code:
                    codes[name.lower()] = code

    sheet = wb["Entries"] if "Entries" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        print(json.dumps({"entries": [], "codes": codes}))
        return

    headers = [norm(h).lower() for h in rows[0]]
    day_keys = {
        "monday": "monday",
        "tuesday": "tuesday",
        "wednesday": "wednesday",
        "thursday": "thursday",
        "friday": "friday",
        "saturday": "saturday",
        "sunday": "sunday",
        "mon": "monday",
        "tue": "tuesday",
        "wed": "wednesday",
        "thu": "thursday",
        "fri": "friday",
        "sat": "saturday",
        "sun": "sunday",
    }
    for row in rows[1:]:
        if not any(v is not None and str(v).strip() for v in row):
            continue
        item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        project_name = norm(item.get("project_name") or item.get("activity") or item.get("code_name") or item.get("engagement"))
        internal_order_code = norm(item.get("internal_order_code") or item.get("code") or item.get("internal order"))
        if not internal_order_code and project_name:
            internal_order_code = codes.get(project_name.lower(), "")
        hours_by_day = {}
        for header, value in item.items():
            key = day_keys.get(header)
            if key and value not in (None, ""):
                try:
                    hours_by_day[key] = float(value)
                except Exception:
                    hours_by_day[key] = value
        total = item.get("hours") or item.get("daily_hours")
        if total in (None, "") and hours_by_day:
            numeric = [v for v in hours_by_day.values() if isinstance(v, (int, float))]
            total = sum(numeric) if numeric else None
        entries.append(
            {
                "project_name": project_name,
                "internal_order_code": internal_order_code,
                "project_type": norm(item.get("project_type") or item.get("group") or "Internal Order"),
                "hours": total,
                "hours_by_day": hours_by_day,
            }
        )

    print(json.dumps({"entries": entries, "codes": codes}, ensure_ascii=False))


if __name__ == "__main__":
    main()
